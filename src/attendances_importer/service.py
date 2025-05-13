import csv
import json
from io import BytesIO, StringIO
from typing import Any

import httpx
import orjson
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError

from .config import Settings
from .dtos import AttendanceCreate, AttendanceCreateMultiple, AttendanceImportResponse
from .errors import NoActiveSheet, RowValidationError

BASE_FIELDS = [
    'full_name',
    'document',
    'document_type',
    'gender',
    'address',
    'birth_date',
    'reason',
]


def read_attendances(reader):
    headers = next(reader)
    additional_keys = headers[len(BASE_FIELDS) :]
    attendance_list = []

    for i, row in enumerate(reader):
        if i >= Settings.MAX_ROWS:
            break

        if not any(row):
            continue

        base_data: dict[str, Any] = dict(zip(BASE_FIELDS, row[: len(BASE_FIELDS)]))
        base_data['document'] = str(base_data['document'])
        additional_values = row[len(BASE_FIELDS) :]
        additional_data = (
            dict(zip(additional_keys, additional_values)) if additional_keys else None
        )
        base_data['additional_data'] = additional_data

        try:
            attendance_list.append(AttendanceCreate(**base_data))
        except ValidationError as e:
            msg = e.errors()[0]['msg']
            loc = e.errors()[0]['loc']
            raise RowValidationError(i + 2, f'{loc[-1]} - {msg}')

    return attendance_list


def read_attendances_from_excel(file: UploadFile):
    contents = file.file.read()
    wb = load_workbook(filename=BytesIO(contents))
    sheet = wb.active
    if not sheet:
        raise NoActiveSheet

    reader = sheet.iter_rows(values_only=True, max_row=Settings.MAX_ROWS)
    return read_attendances(reader)


def read_attendances_from_csv(file: UploadFile):
    contents = file.file.read().decode('utf-8')
    reader = csv.reader(StringIO(contents))
    return read_attendances(reader)


async def create_attendances(access_token: str, attendances: list[AttendanceCreate]):
    content = orjson.dumps(
        AttendanceCreateMultiple(attendances=attendances).model_dump()
    )
    async with httpx.AsyncClient() as client:
        response = await client.post(
            f'{Settings.ATTENDANCES_URL}/attendances/multiple',
            headers={'Authorization': 'Bearer ' + access_token},
            content=content,
        )
        if response.is_error:
            try:
                detail = response.json()['detail']
            except (json.decoder.JSONDecodeError, KeyError):
                detail = response.text
            raise HTTPException(
                status_code=response.status_code,
                detail=detail,
            )

        return response


async def from_excel(access_token: str, file: UploadFile):
    attendances = read_attendances_from_excel(file)
    response = await create_attendances(access_token, attendances)
    return AttendanceImportResponse(
        attendances=attendances,
        insertion_response=response.json(),
        file_extension='xlsx',
    )


async def from_csv(access_token: str, file: UploadFile):
    attendances = read_attendances_from_csv(file)
    response = await create_attendances(access_token, attendances)
    return AttendanceImportResponse(
        attendances=attendances,
        insertion_response=response.json(),
        file_extension='csv',
    )
